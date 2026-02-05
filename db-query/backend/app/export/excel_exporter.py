"""Excel exporter implementation."""

import io
import time
from typing import Any, BinaryIO, AsyncIterator
from openpyxl import Workbook
from openpyxl.styles import Font
from app.export.base import BaseExporter, ExportOptions, ExportResult


class ExcelExporter(BaseExporter):
    """Excel exporter - uses openpyxl."""

    async def export(
        self,
        columns: list[dict[str, str]],
        rows: list[dict[str, Any]],
        options: ExportOptions,
    ) -> ExportResult:
        """Export data to Excel format."""
        start_time = time.time()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = options.sheet_name

        # Write headers
        if options.include_headers:
            headers = [col["name"] for col in columns]
            ws.append(headers)

            # Bold headers
            for cell in ws[1]:
                cell.font = Font(bold=True)

        # Write data
        for row in rows:
            row_data = [row.get(col["name"]) for col in columns]
            ws.append(row_data)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            # Set width with padding, max 50
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        file_data = output.getvalue()

        return ExportResult(
            file_data=file_data,
            file_name=self._generate_filename(self.get_file_extension()),
            mime_type=self.get_mime_type(),
            row_count=len(rows),
            file_size_bytes=len(file_data),
            export_time_ms=int((time.time() - start_time) * 1000),
        )

    def get_file_extension(self) -> str:
        """Return 'xlsx'."""
        return "xlsx"

    def get_mime_type(self) -> str:
        """Return Excel MIME type."""
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def supports_streaming(self) -> bool:
        """Excel supports streaming with openpyxl."""
        return True

    async def stream_export(
        self,
        columns: list[dict[str, str]],
        row_iterator: AsyncIterator[dict[str, Any]],
        output: BinaryIO,
        options: ExportOptions,
    ) -> int:
        """Stream Excel export for large datasets."""
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = options.sheet_name

        # Write headers
        if options.include_headers:
            headers = [col["name"] for col in columns]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)

        # Write rows
        row_count = 0
        async for row in row_iterator:
            row_data = [row.get(col["name"]) for col in columns]
            ws.append(row_data)
            row_count += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to output
        wb.save(output)
        return row_count
